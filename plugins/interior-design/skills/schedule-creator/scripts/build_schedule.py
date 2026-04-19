#!/usr/bin/env python3
"""
build_schedule.py — Interior Design Room Schedule Excel Generator

Always produces THREE sheets:
  1. "[Project] Schedule"    — all rooms with blue section headers per room
  2. "Materials & Finishes"  — finish materials grouped by type
  3. "Room Finish Schedule"  — technical CAD/Revit-compatible table: one row
                               per room showing which finish applies to each
                               surface (floor, base, N/E/S/W walls, ceiling)

Usage:
  python3 build_schedule.py \
    --output "/path/to/Kitchen Schedule.xlsx" \
    --data   "/path/to/products.json" \
    [--project "Intersecting Stories"] \
    [--studio  "ARCH X482.2 — Design Studio II"] \
    [--semester "SPRING 2026"]

See SKILL.md for the full JSON schema.
"""

import argparse
import json
import os
import shutil
import tempfile
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.utils.units import pixels_to_EMU
    from PIL import Image as PILImage
    import requests
except ImportError as e:
    print(f"Missing dependency: {e}")
    print("Install with: pip install openpyxl pillow requests")
    raise

# ── Layout constants (main schedule) ─────────────────────────────────────────
COL_WIDTHS    = {'A': 10, 'B': 17, 'C': 28, 'D': 20, 'E': 30, 'F': 40, 'G': 10, 'H': 30}
HDR_ROW_H     = 22
SECTION_ROW_H = 22
CAT_ROW_H     = 18
PROD_ROW_H    = 90
IMG_MAX_W     = 110
IMG_MAX_H     = 78
COL_B_PX      = 128
ROW_PROD_PX   = 120

# ── Room Finish Schedule layout ───────────────────────────────────────────────
RFS_COL_WIDTHS = {
    'A': 9,   # ROOM NO.
    'B': 22,  # ROOM NAME
    'C': 14,  # FLOOR
    'D': 12,  # BASE
    'E': 14,  # NORTH
    'F': 14,  # EAST
    'G': 14,  # SOUTH
    'H': 14,  # WEST
    'I': 10,  # HEIGHT
    'J': 26,  # CEILING MATERIAL / FINISH
    'K': 28,  # NOTES
}
RFS_HEADERS = [
    'ROOM NO.', 'ROOM NAME', 'FLOOR', 'BASE',
    'NORTH', 'EAST', 'SOUTH', 'WEST',
    'HEIGHT', 'CEILING MATERIAL / FINISH', 'NOTES',
]
RFS_HDR_H   = 30
RFS_DATA_H  = 18

# ── Materials & finishes routing ──────────────────────────────────────────────
FINISH_PREFIX_MAP = {
    'PT':  'PAINT',
    'BB':  'MOULDING',
    'HW':  'HARDWOOD',
    'FT':  'TILE',
    'WT':  'TILE',
    'WC':  'WALLCOVERING',
    'CT':  'STONE',
    'ST':  'STONE',
    'MT':  'METAL',
}
FINISH_CATEGORY_ORDER = ['PAINT', 'MOULDING', 'HARDWOOD', 'TILE', 'WALLCOVERING', 'STONE', 'METAL']


def _code_prefix(code):
    return code.strip().split('-')[0].upper()


def is_finish(code):
    return _code_prefix(code) in FINISH_PREFIX_MAP


def finish_category(code):
    return FINISH_PREFIX_MAP.get(_code_prefix(code), 'OTHER')


def normalize_stone_code(code, stone_counter):
    if _code_prefix(code) == 'CT':
        stone_counter['n'] += 1
        return f"ST-{stone_counter['n']}"
    return code


# ── Style helpers (main schedule) ─────────────────────────────────────────────
def _thin():  return Side(style='thin',   color='CCCCCC')
def _med():   return Side(style='medium', color='000000')


def _hdr_border():
    return Border(top=_med(), bottom=_med(), left=_med(), right=_med())


def _item_border(col):
    left  = _med() if col == 1 else _thin()
    right = _med() if col == 8 else _thin()
    return Border(top=_thin(), bottom=_thin(), left=left, right=right)


def _fill(hex_color):    return PatternFill('solid', fgColor=hex_color)
def _bold(sz=9):         return Font(name='Arial', bold=True,  size=sz)
def _reg(sz=8):          return Font(name='Arial', bold=False, size=sz)
def _link_font(sz=11):   return Font(name='Arial', size=sz)
def _section_font():     return Font(name='Arial', bold=True,  size=10, color='FFFFFFFF')
def _center(wrap=True):  return Alignment(horizontal='center', vertical='center', wrap_text=wrap)
def _left(wrap=False):   return Alignment(horizontal='left',   vertical='center', wrap_text=wrap)


# ── Style helpers (Room Finish Schedule — plain / CAD-compatible) ─────────────
def _rfs_thin():   return Side(style='thin',   color='000000')
def _rfs_med():    return Side(style='medium', color='000000')


def _rfs_outer_border(row, total_rows, col, total_cols):
    """Medium border on the outer edge of the table, thin inside."""
    top    = _rfs_med() if row == 1          else _rfs_thin()
    bottom = _rfs_med() if row == total_rows else _rfs_thin()
    left   = _rfs_med() if col == 1          else _rfs_thin()
    right  = _rfs_med() if col == total_cols else _rfs_thin()
    return Border(top=top, bottom=bottom, left=left, right=right)


def _rfs_bold(sz=9):  return Font(name='Arial', bold=True,  size=sz)
def _rfs_reg(sz=9):   return Font(name='Arial', bold=False, size=sz)
def _rfs_center():    return Alignment(horizontal='center', vertical='center', wrap_text=True)
def _rfs_left():      return Alignment(horizontal='left',   vertical='center', wrap_text=True)


# ── Image utilities ───────────────────────────────────────────────────────────
def download_image(url, dest_path):
    if not url or not str(url).startswith('http'):
        return False
    agents = [
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/120.0.0.0',
        'curl/7.68.0', '',
    ]
    for agent in agents:
        try:
            headers = {'User-Agent': agent} if agent else {}
            r = requests.get(url, headers=headers, timeout=20, stream=True)
            if r.status_code == 200:
                content = b''.join(r.iter_content(8192))
                if len(content) > 2000:
                    with open(dest_path, 'wb') as f:
                        f.write(content)
                    if os.path.getsize(dest_path) > 2000:
                        return True
        except Exception:
            continue
    return False


def make_thumbnail(src_path, tmp_dir, max_w=IMG_MAX_W, max_h=IMG_MAX_H):
    if not src_path or not os.path.exists(str(src_path)):
        return None
    try:
        pil = PILImage.open(str(src_path)).convert('RGB')
        pil.thumbnail((max_w, max_h), PILImage.LANCZOS)
        name = 'thumb_' + Path(src_path).name.replace(' ', '_') + '.png'
        dst = Path(tmp_dir) / name
        pil.save(str(dst), 'PNG')
        return str(dst), pil.width, pil.height
    except Exception as e:
        print(f'  Thumbnail error {src_path}: {e}')
        return None


def embed_image(ws, img_path, img_w, img_h, col_idx, row_idx):
    try:
        xl_img        = XLImage(img_path)
        xl_img.width  = img_w
        xl_img.height = img_h
        x_off = max(0, (COL_B_PX    - img_w) // 2)
        y_off = max(0, (ROW_PROD_PX - img_h) // 2)
        marker = AnchorMarker(col=col_idx, colOff=pixels_to_EMU(x_off),
                              row=row_idx,  rowOff=pixels_to_EMU(y_off))
        size = XDRPositiveSize2D(pixels_to_EMU(img_w), pixels_to_EMU(img_h))
        xl_img.anchor = OneCellAnchor(_from=marker, ext=size)
        ws.add_image(xl_img)
        return True
    except Exception as e:
        print(f'  Image embed error: {e}')
        return False


# ── Sheet setup helpers ───────────────────────────────────────────────────────
def _setup_sheet(ws):
    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows       = '1:1'


def _write_column_headers(ws, row):
    ws.row_dimensions[row].height = HDR_ROW_H
    headers = ['CODE', 'IMAGE', 'MODEL', 'MANUFACTURER', 'DIMENSIONS', 'NOTES', 'QUANTITY', 'LINK']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font      = _bold(9)
        c.fill      = _fill('E0E0E0')
        c.alignment = _center(wrap=False)
        c.border    = _hdr_border()
    ws.freeze_panes = f'A{row + 1}'


def _write_section_header(ws, row, text):
    ws.row_dimensions[row].height = SECTION_ROW_H
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = _section_font()
    c.fill      = _fill('2C5F8A')
    c.alignment = _left()
    c.border    = _hdr_border()


def _write_category_header(ws, row, text):
    ws.row_dimensions[row].height = CAT_ROW_H
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = _bold(9)
    c.fill      = _fill('D6E4F0')
    c.alignment = _left()
    c.border    = _hdr_border()


def _write_item_row(ws, row, item, item_idx, tmp_dir, cache_dir):
    fill = _fill('F4F4F4' if item_idx % 2 == 1 else 'FFFFFF')
    ws.row_dimensions[row].height = PROD_ROW_H

    def write(col, val='', fnt=None, hyper=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = fnt or _reg()
        c.alignment = _center()
        c.fill      = fill
        c.border    = _item_border(col)
        if hyper:
            c.hyperlink = hyper
        return c

    write(1, item.get('code', ''), fnt=_bold(9))
    write(2, '')
    write(3, item.get('model', ''))
    write(4, item.get('mfr', ''))
    write(5, item.get('dims', ''))
    write(6, item.get('notes', ''))
    write(7, item.get('qty', ''))
    link_val = item.get('link', '')
    if link_val:
        write(8, 'LINK', fnt=_link_font(), hyper=link_val)
    else:
        write(8, '')

    img_src = item.get('img')
    if img_src:
        if str(img_src).startswith('http'):
            if cache_dir:
                slug   = item.get('code', 'img').replace('-', '_').lower()
                cached = Path(cache_dir) / f'{slug}.jpg'
                if not (cached.exists() and cached.stat().st_size > 2000):
                    download_image(img_src, str(cached))
                img_src = str(cached) if cached.exists() else None
            else:
                dl_path = Path(tmp_dir) / f'dl_{row}.jpg'
                download_image(img_src, str(dl_path))
                img_src = str(dl_path) if dl_path.exists() else None
        if img_src and os.path.exists(img_src):
            result = make_thumbnail(img_src, tmp_dir)
            if result:
                thumb_path, img_w, img_h = result
                return embed_image(ws, thumb_path, img_w, img_h,
                                   col_idx=1, row_idx=row - 1)
    return False


# ── Room Finish Schedule sheet writer ────────────────────────────────────────
def _fmt_surface(codes):
    """Convert a surface assignment to display string. Multiple codes joined with ' / '."""
    if not codes:
        return ''
    if isinstance(codes, str):
        return codes.strip()
    return ' / '.join(str(c).strip() for c in codes if c)


def write_room_finish_sheet(wb, rooms, project='', studio='', semester=''):
    """
    Create (or recreate) the 'Room Finish Schedule' sheet in wb.

    Each room dict may carry:
      room_no        (str|int)  — overrides auto-numbering if provided
      ceiling_height (str)      — e.g. "9'-0\""
      ceiling_finish (str)      — e.g. "GYP. BD. / PT-1"
      rfs_notes      (str)      — row-level notes
      surfaces       (dict)     — keys: floor, base, north, east, south, west
                                   values: str or list[str] of finish codes

    Rooms are auto-numbered 101, 102, … unless room_no is provided.
    Plain CAD-compatible styling: no fill colors, Arial font, black borders.
    """
    if 'Room Finish Schedule' in wb.sheetnames:
        del wb['Room Finish Schedule']

    ws = wb.create_sheet('Room Finish Schedule')

    for col, w in RFS_COL_WIDTHS.items():
        ws.column_dimensions[col].width = w

    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows       = '1:1'

    n_cols      = len(RFS_HEADERS)
    total_rows  = 1 + len(rooms)

    # Header row
    ws.row_dimensions[1].height = RFS_HDR_H
    for ci, h in enumerate(RFS_HEADERS, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font      = _rfs_bold(9)
        c.alignment = _rfs_center()
        c.border    = _rfs_outer_border(1, total_rows, ci, n_cols)
        c.fill      = PatternFill('solid', fgColor='F2F2F2')

    ws.freeze_panes = 'A2'

    # Data rows
    for i, room in enumerate(rooms):
        data_row = i + 2
        ws.row_dimensions[data_row].height = RFS_DATA_H

        auto_no = 100 + (i + 1)
        room_no = room.get('room_no', str(auto_no))
        surfs   = room.get('surfaces', {})

        values = [
            str(room_no),
            room.get('name', ''),
            _fmt_surface(surfs.get('floor')),
            _fmt_surface(surfs.get('base')),
            _fmt_surface(surfs.get('north')),
            _fmt_surface(surfs.get('east')),
            _fmt_surface(surfs.get('south')),
            _fmt_surface(surfs.get('west')),
            room.get('ceiling_height', ''),
            room.get('ceiling_finish', ''),
            room.get('rfs_notes', ''),
        ]

        for ci, val in enumerate(values, 1):
            c = ws.cell(row=data_row, column=ci, value=val)
            c.font      = _rfs_reg(9)
            c.alignment = _rfs_center() if ci == 1 else _rfs_left()
            c.border    = _rfs_outer_border(data_row, total_rows, ci, n_cols)

    hdr = 'ROOM FINISHES SCHEDULE'
    if project:
        hdr += f' — {project.upper()}'
    ws.oddHeader.center.text = hdr
    footer_left = studio + (f' — {semester}' if semester else '')
    ws.oddFooter.left.text  = footer_left
    ws.oddFooter.right.text = 'RESIDENTIAL INTERIOR DESIGN'

    print(f'  Sheet 3 "Room Finish Schedule": {len(rooms)} room rows')
    return ws


# ── Main build function ───────────────────────────────────────────────────────
def build_schedule(rooms, output_path,
                   project='', studio='', semester='',
                   cache_dir=None):
    """
    Build the full room schedule workbook (3 sheets).

    rooms: list of room dicts. Each room:
      {
        'name':           'Kitchen',
        'room_no':        '101',
        'ceiling_height': "9'-0\"",
        'ceiling_finish': 'PT-1',
        'rfs_notes':      '',
        'surfaces': {
            'floor':  ['FT-1'],
            'base':   ['BB-1'],
            'north':  ['PT-1'],
            'east':   ['WC-1', 'PT-1'],
            'south':  ['PT-1'],
            'west':   ['PT-1'],
        },
        'categories': [
          {
            'label': 'APPLIANCES',
            'items': [{'code': 'AP-1', 'img': '...', 'model': '...', ...}]
          }
        ]
      }
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    tmp_dir = tempfile.mkdtemp()
    if cache_dir:
        Path(cache_dir).mkdir(parents=True, exist_ok=True)

    try:
        wb = openpyxl.Workbook()

        # Sheet 1: Main schedule
        schedule_title = f'{project} Schedule' if project else 'Room Schedule'
        ws_main = wb.active
        ws_main.title = schedule_title
        _setup_sheet(ws_main)

        # Sheet 2: Materials & Finishes
        ws_mf = wb.create_sheet('Materials & Finishes')
        _setup_sheet(ws_mf)

        # Separate items: main vs. finish materials
        main_rooms    = []
        finish_groups = defaultdict(list)

        for room in rooms:
            room_cats = []
            for cat in room.get('categories', []):
                keep_items = []
                for item in cat.get('items', []):
                    code = item.get('code', '')
                    if is_finish(code):
                        finish_groups[finish_category(code)].append(item)
                    else:
                        keep_items.append(item)
                if keep_items:
                    room_cats.append({'label': cat['label'], 'items': keep_items})
            if room_cats:
                main_rooms.append({'name': room['name'], 'categories': room_cats})

        # Write Sheet 1
        _write_column_headers(ws_main, 1)
        cur_row   = 2
        item_idx  = 0
        img_count = 0

        for room in main_rooms:
            _write_section_header(ws_main, cur_row, room['name'].upper())
            cur_row += 1
            for cat in room['categories']:
                _write_category_header(ws_main, cur_row, cat['label'])
                cur_row += 1
                for item in cat['items']:
                    item_idx += 1
                    if _write_item_row(ws_main, cur_row, item, item_idx, tmp_dir, cache_dir):
                        img_count += 1
                    cur_row += 1

        room_names = ' + '.join(r['name'] for r in rooms)
        hdr = room_names.upper() + ' SCHEDULE'
        if project:
            hdr += f' — {project.upper()}'
        footer_left = studio + (f' — {semester}' if semester else '')
        ws_main.oddHeader.center.text = hdr
        ws_main.oddFooter.left.text   = footer_left
        ws_main.oddFooter.right.text  = 'RESIDENTIAL INTERIOR DESIGN'
        print(f'  Sheet 1 "{schedule_title}": {cur_row - 2} rows, {img_count} images')

        # Write Sheet 2
        _write_column_headers(ws_mf, 1)
        mf_row      = 2
        mf_item_idx = 0
        mf_img      = 0
        _write_section_header(ws_mf, mf_row, 'MATERIALS & FINISHES')
        mf_row += 1

        stone_counter = {'n': 0}
        ordered_keys  = FINISH_CATEGORY_ORDER + [
            k for k in finish_groups if k not in FINISH_CATEGORY_ORDER
        ]
        for cat_name in ordered_keys:
            items = finish_groups.get(cat_name)
            if not items:
                continue
            _write_category_header(ws_mf, mf_row, cat_name)
            mf_row += 1
            for item in items:
                mf_item_idx += 1
                display_item = dict(item)
                display_item['code'] = normalize_stone_code(item.get('code', ''), stone_counter)
                if _write_item_row(ws_mf, mf_row, display_item, mf_item_idx, tmp_dir, cache_dir):
                    mf_img += 1
                mf_row += 1

        ws_mf.oddHeader.center.text = (
            f'MATERIALS & FINISHES — {project.upper()}' if project else 'MATERIALS & FINISHES'
        )
        ws_mf.oddFooter.left.text  = footer_left
        ws_mf.oddFooter.right.text = 'RESIDENTIAL INTERIOR DESIGN'
        print(f'  Sheet 2 "Materials & Finishes": {mf_row - 2} rows, {mf_img} images')

        # Write Sheet 3
        write_room_finish_sheet(wb, rooms, project=project,
                                studio=studio, semester=semester)

        wb.save(str(output_path))
        size = output_path.stat().st_size
        print(f'✅ Schedule saved: {output_path}  ({size // 1024} KB)')
        return str(output_path)

    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


# ── CLI entry point ───────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description='Build interior design room schedule Excel.')
    parser.add_argument('--output',    required=True)
    parser.add_argument('--data',      required=True)
    parser.add_argument('--project',   default='')
    parser.add_argument('--studio',    default='')
    parser.add_argument('--semester',  default='')
    parser.add_argument('--cache-dir', default=None)
    args = parser.parse_args()

    with open(args.data) as f:
        data = json.load(f)

    if 'rooms' in data:
        rooms = data['rooms']
    elif 'categories' in data:
        rooms = [{'name': data.get('room', 'Room'), 'categories': data['categories']}]
    else:
        raise ValueError("JSON must contain 'rooms' (multi-room) or 'categories' (single-room)")

    build_schedule(
        rooms       = rooms,
        output_path = args.output,
        project     = args.project  or data.get('project', ''),
        studio      = args.studio   or data.get('studio', ''),
        semester    = args.semester or data.get('semester', ''),
        cache_dir   = args.cache_dir,
    )


if __name__ == '__main__':
    main()
