#!/usr/bin/env python3
"""
annotate_spec_sheets.py
=======================
Annotate spec sheet PDFs using codes from a room/FF&E schedule spreadsheet.

For each product in the schedule:
  - Find the corresponding spec sheet pages (by matching model numbers)
  - Stamp each page with the schedule code in red at top-right
  - Yellow-highlight installation-critical text (dimensions, clearances,
    electrical specs, duct sizes, etc.)

Then combine all annotated pages into a single output PDF in schedule order.

Usage
-----
python3 annotate_spec_sheets.py \
    --schedule  "Kitchen Schedule.xlsx" \
    --spec-dir  "Kitchen/Spec Sheets" \
    --output    "Kitchen Schedule - Spec Sheets.pdf"

Optional:
    --schedule-sheet  SHEET_NAME   (default: first sheet)
    --order           schedule|original  (default: schedule)
    --manual-map      JSON_STRING  (override page assignments, see --help)
"""

import argparse
import json
import re
import sys
from pathlib import Path
from collections import defaultdict, OrderedDict

try:
    import fitz  # PyMuPDF
except ImportError:
    sys.exit("ERROR: PyMuPDF not installed. Run: pip install pymupdf --break-system-packages")

try:
    import openpyxl
except ImportError:
    sys.exit("ERROR: openpyxl not installed. Run: pip install openpyxl --break-system-packages")


# ── Installation-critical highlight patterns ─────────────────────────────────
# These match text lines that contain info a contractor / installer must follow.
HIGHLIGHT_PATTERNS = [
    # Overall / product / installation dimensions
    r"overall\s+appliance\s+dimension",
    r"overall\s+dimension",
    r"appliance\s+dimension",
    r"product\s+dimension",
    r"installation\s+dimension",
    r"specification",
    # Cutout / rough-in / opening
    r"required\s+cutout",
    r"cutout\s+size",
    r"cut.?out",
    r"rough.?in",
    r"opening\s+size",
    r"cabinet\s+opening",
    r"cabinet\s+size",
    r"min\.?\s+cabinet",
    # Dimension notation patterns
    r"hxwxd",
    r"wxdxh",
    r"lxwxh",
    r"h\s*[×x]\s*w\s*[×x]\s*d",
    r"dimension.{0,8}(in\.|mm|inches|cm)",
    # Individual dimension rows (Depth / Height / Width as labeled fields)
    r"^\s*depth\s",
    r"^\s*height\s",
    r"^\s*width\s",
    r"overall\s+(height|width|depth)",
    r"(height|width|depth)\s*[\:\-–]?\s*\d",
    r"adjustable\s+range\s+height",
    r"niche\s+(depth|height|width)",
    # Clearance requirements
    r"clearance",
    r"minimum\s+clearance",
    r"required\s+clearance",
    r"minimum\s+space",
    r"side\s+(clearance|spacing)",
    r"rear\s+clearance",
    r"overhead\s+clearance",
    r"min\.?\s+(4|6|12|18|24)\s*[\"'″]",  # min XX" clearance
    # Electrical / circuit
    r"circuit\s+breaker",
    r"required\s+voltage",
    r"supply\s+voltage",
    r"plug\s+type",
    r"connection\s+type",
    r"^.{0,35}connection\s*[\:\-]",
    r"^.{0,35}service\s*[\:\-]",
    r"\b120\s*v\b",
    r"\b208\s*v\b",
    r"\b240\s*v\b",
    r"\b15\s*a\b",
    r"\b20\s*a\b",
    r"\b40\s*a\b",
    r"\bvolts?\b",
    r"\bampere",
    r"electrical\s+connection",
    r"power\s+(supply|requirement|cord|connection)",
    r"nema\s+\d",
    # Ducting / ventilation
    r"duct\s+(size|diameter|opening|connection)",
    r"vent(ilation)?\s+(size|opening|duct)",
    r"duct(ing)?\s+size",
    r"round\s+duct",
    # Weight / structural
    r"net\s+weight",
    r"gross\s+weight",
    r"weight\s*\(?lbs",
    r"weight\s*\(?kg",
    # Plumbing / water
    r"water\s+supply",
    r"supply\s+line",
    r"drain\s+(connection|size|opening)",
    r"hole\s+size",
    r"hole\s+diameter",
]

COMPILED_PATTERNS = [re.compile(p, re.IGNORECASE | re.MULTILINE) for p in HIGHLIGHT_PATTERNS]


def line_is_install_critical(line_text: str) -> bool:
    """Return True if this text line contains installation-critical information."""
    for pattern in COMPILED_PATTERNS:
        if pattern.search(line_text):
            return True
    return False


# ── Schedule reading ──────────────────────────────────────────────────────────

def read_schedule(xlsx_path: str, sheet_name: str = None) -> list[tuple[str, list[str]]]:
    """
    Read a room/FF&E schedule and return an ordered list of (code, [model_tokens]).

    The schedule is expected to have a CODE column and a MODEL column.
    Category header rows (e.g., 'APPLIANCES', 'PLUMBING FIXTURES') and rows
    with no model number are skipped.

    Returns a list like:
      [("AP-1", ["OR30SCI6X1"]), ("AP-2", ["RS36W80RJ1", "RS36W80RJ1N"]), ...]
    Multiple model tokens allow partial matching against PDFs.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    # Find header row
    code_col = model_col = None
    header_row = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and str(cell.value).strip().upper() == "CODE":
                code_col = cell.column
                header_row = cell.row
            if cell.value and str(cell.value).strip().upper() == "MODEL":
                model_col = cell.column
        if code_col and model_col:
            break

    if not code_col or not model_col:
        # Fallback: assume first two columns are CODE and MODEL
        code_col, model_col, header_row = 1, 3, 1
        print("WARNING: Could not find CODE/MODEL headers; defaulting to columns 1 and 3.")

    entries = []  # (code, [model_tokens])
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        code_val = row[code_col - 1] if len(row) >= code_col else None
        model_val = row[model_col - 1] if len(row) >= model_col else None

        if not code_val or not model_val:
            continue
        code_str = str(code_val).strip()
        model_str = str(model_val).strip()

        # Skip category headers (no hyphen in code, e.g. "APPLIANCES")
        if "-" not in code_str:
            continue
        # Skip if model looks like a header
        if model_str.upper() == model_str and len(model_str) > 20:
            continue

        # Build a list of search tokens from the model string.
        # Split on spaces, commas, slashes, AND hyphens so that finish-code
        # suffixes like "-SS" in "64243LF-SS" don't prevent matching "64243LF"
        # in the spec PDF.  Keep only tokens ≥5 chars to avoid false positives
        # on short fragments like "SS", "24", etc.
        tokens = [model_str]
        for token in re.split(r"[\s,/\-]+", model_str):
            token = token.strip()
            if len(token) >= 5:
                tokens.append(token)
        # Deduplicate while preserving order
        seen = set()
        unique_tokens = []
        for t in tokens:
            if t not in seen:
                seen.add(t)
                unique_tokens.append(t)

        entries.append((code_str, unique_tokens))
        print(f"  Schedule entry: {code_str} → {model_str}")

    return entries


# ── PDF page → code mapping ───────────────────────────────────────────────────

def is_already_annotated(pdf_path: str) -> bool:
    """
    Return True if this PDF already has highlight annotations — a strong
    signal it was produced by a previous run of this script and should be
    skipped to avoid processing the output as if it were a source file.
    """
    try:
        doc = fitz.open(pdf_path)
        for page in doc:
            if list(page.annots(types=[fitz.PDF_ANNOT_HIGHLIGHT])):
                doc.close()
                return True
        doc.close()
    except Exception:
        pass
    return False


def build_page_code_map(
    pdf_paths: list[str],
    schedule_entries: list[tuple[str, list[str]]],
    manual_map: dict = None,
) -> dict[tuple[str, int], str]:
    """
    For each (pdf_path, page_index), determine the schedule code.

    Strategy: "section detection"
    - Scan every page for model number tokens from the schedule.
    - The first page that matches a given code starts a new section.
    - All subsequent pages until the next section match inherit that code.
    - Pages before any match are unassigned.

    manual_map format:
      {"AP-4": [{"file": "bosch_spec.pdf", "pages": [0, 1, 2]}]}
    """
    page_code: dict[tuple[str, int], str] = {}

    # Build a flat lookup: token (lowercase) → code
    # Longer tokens take priority to avoid spurious short matches.
    token_to_code: list[tuple[str, str]] = []
    for code, tokens in schedule_entries:
        for t in tokens:
            token_to_code.append((t.lower(), code))
    # Sort by token length descending so longer (more specific) tokens match first
    token_to_code.sort(key=lambda x: -len(x[0]))

    def find_code_in_text(text: str) -> str | None:
        text_lower = text.lower()
        for token, code in token_to_code:
            if token in text_lower:
                return code
        return None

    for pdf_path in pdf_paths:
        try:
            doc = fitz.open(pdf_path)
        except Exception as e:
            print(f"WARNING: Could not open {pdf_path}: {e}")
            continue

        current_code = None
        fname = Path(pdf_path).name

        for page_idx in range(len(doc)):
            page = doc[page_idx]
            text = page.get_text()

            detected = find_code_in_text(text)
            if detected:
                current_code = detected

            if current_code:
                page_code[(pdf_path, page_idx)] = current_code

        doc.close()

    # Apply manual overrides
    if manual_map:
        for code, file_entries in manual_map.items():
            for entry in file_entries:
                target_file = entry.get("file", "")
                for pdf_path in pdf_paths:
                    if Path(pdf_path).name == target_file or pdf_path == target_file:
                        for pg in entry.get("pages", []):
                            page_code[(pdf_path, pg)] = code

    # Print assignment summary
    code_counts: dict[str, int] = defaultdict(int)
    for v in page_code.values():
        code_counts[v] += 1
    unassigned = sum(
        1 for pdf_path in pdf_paths
        for pg_idx in range(fitz.open(pdf_path).page_count)
        if (pdf_path, pg_idx) not in page_code
    )
    print("\nPage assignment summary:")
    for code, count in sorted(code_counts.items()):
        print(f"  {code}: {count} page(s)")
    if unassigned:
        print(f"  (unassigned: {unassigned} page(s) — no matching model found)")

    return page_code


# ── Annotation ────────────────────────────────────────────────────────────────

def add_code_label(page: fitz.Page, code: str):
    """Stamp the schedule code in red at the top-right corner of the page."""
    font_size = 16
    padding = 10
    approx_char_w = font_size * 0.55
    text_w = len(code) * approx_char_w + 6
    text_h = font_size + 4

    x1 = page.rect.width - padding
    y_baseline = padding + font_size

    # White background for legibility over any existing content
    bg = fitz.Rect(x1 - text_w - 4, padding - 2, x1 + 2, padding + text_h + 2)
    shape = page.new_shape()
    shape.draw_rect(bg)
    shape.finish(color=(1, 1, 1), fill=(1, 1, 1), width=0)
    shape.commit()

    # Red text
    page.insert_text(
        fitz.Point(x1 - text_w, y_baseline),
        code,
        fontsize=font_size,
        fontname="helv",
        color=(0.85, 0, 0),
    )


def highlight_install_critical(page: fitz.Page):
    """Find and yellow-highlight installation-critical lines on this page."""
    try:
        blocks = page.get_text("dict", flags=fitz.TEXT_PRESERVE_WHITESPACE)["blocks"]
    except Exception:
        return

    for block in blocks:
        if block.get("type") != 0:  # skip non-text blocks
            continue
        for line in block.get("lines", []):
            line_text = " ".join(span.get("text", "") for span in line.get("spans", []))
            if not line_text.strip():
                continue
            if line_is_install_critical(line_text):
                bbox = line.get("bbox")
                if not bbox:
                    continue
                rect = fitz.Rect(bbox[0] - 1, bbox[1] - 1, bbox[2] + 1, bbox[3] + 1)
                try:
                    annot = page.add_highlight_annot(rect)
                    annot.set_colors(stroke=(1, 1, 0))
                    annot.update()
                except Exception:
                    pass


def annotate_page(page: fitz.Page, code: str):
    """Apply all annotations to a single page."""
    highlight_install_critical(page)
    add_code_label(page, code)


# ── Assembly ──────────────────────────────────────────────────────────────────

def build_output_pdf(
    pdf_paths: list[str],
    page_code_map: dict[tuple[str, int], str],
    schedule_entries: list[tuple[str, list[str]]],
    order: str = "schedule",
    output_path: str = "annotated_spec_sheets.pdf",
):
    """
    Annotate pages and write the combined output PDF.

    order='schedule'  → pages grouped by code in schedule order (AP-1 first, etc.)
    order='original'  → pages in their original PDF order, with annotations
    """
    # Open all source docs
    open_docs: dict[str, fitz.Document] = {}
    for path in pdf_paths:
        try:
            open_docs[path] = fitz.open(path)
        except Exception as e:
            print(f"WARNING: Cannot open {path}: {e}")

    # Build (pdf_path, page_idx, code) tuples in desired order
    if order == "schedule":
        # Group by code in schedule order
        code_order = [code for code, _ in schedule_entries]
        # Remove duplicates while preserving order
        seen = set()
        unique_code_order = []
        for c in code_order:
            if c not in seen:
                seen.add(c)
                unique_code_order.append(c)

        # Map: code → [(pdf_path, page_idx), ...]
        code_to_pages: dict[str, list] = defaultdict(list)
        for (path, pg_idx), code in page_code_map.items():
            code_to_pages[code].append((path, pg_idx))
        # Sort pages within each code by (pdf_path, page_idx) for stable order
        for code in code_to_pages:
            code_to_pages[code].sort()

        page_tuples = []
        for code in unique_code_order:
            for path, pg_idx in code_to_pages.get(code, []):
                page_tuples.append((path, pg_idx, code))
    else:
        # Original order: iterate PDFs in directory order, pages in sequence
        page_tuples = []
        for path in pdf_paths:
            doc = open_docs.get(path)
            if not doc:
                continue
            for pg_idx in range(len(doc)):
                code = page_code_map.get((path, pg_idx))
                if code:
                    page_tuples.append((path, pg_idx, code))

    if not page_tuples:
        print("ERROR: No pages could be mapped to schedule codes. Check that model "
              "numbers in the schedule match text in the spec PDFs.")
        return

    # Annotate and assemble
    output = fitz.open()
    print(f"\nAssembling {len(page_tuples)} pages…")

    for path, pg_idx, code in page_tuples:
        doc = open_docs.get(path)
        if not doc:
            continue
        # Insert a clean copy of the page, then annotate
        output.insert_pdf(doc, from_page=pg_idx, to_page=pg_idx)
        annotate_page(output[-1], code)

    output.save(output_path, garbage=4, deflate=True)
    print(f"\n✓ Saved: {output_path}")
    print(f"  Total pages: {len(output)}")

    # Print per-code page count
    from collections import Counter
    counts = Counter(code for _, _, code in page_tuples)
    for code in [c for c, _ in schedule_entries if c in counts]:
        print(f"  {code}: {counts[code]} page(s)")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Annotate spec sheet PDFs from a room/FF&E schedule.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--schedule", required=True, help="Path to the room schedule .xlsx")
    parser.add_argument("--spec-dir", required=True,
                        help="Directory containing spec sheet PDF(s)")
    parser.add_argument("--output", required=True, help="Output PDF path")
    parser.add_argument("--schedule-sheet", default=None,
                        help="Sheet name in the xlsx (default: first sheet)")
    parser.add_argument("--order", choices=["schedule", "original"], default="schedule",
                        help="Page output order: 'schedule' (by code) or 'original' (PDF order)")
    parser.add_argument("--manual-map", default=None,
                        help='JSON override for page assignments. '
                             'Example: \'{"AP-4": [{"file": "bosch.pdf", "pages": [0,1,2]}]}\'')
    args = parser.parse_args()

    # Validate paths
    schedule_path = Path(args.schedule)
    spec_dir = Path(args.spec_dir)
    if not schedule_path.exists():
        sys.exit(f"ERROR: Schedule not found: {schedule_path}")
    if not spec_dir.is_dir():
        sys.exit(f"ERROR: Spec directory not found: {spec_dir}")

    # Find all PDFs in spec dir, skipping previously-annotated outputs
    all_pdfs = sorted(str(p) for p in spec_dir.glob("*.pdf"))
    # Also skip the output file itself (in case it already exists in the same dir)
    output_resolved = str(Path(args.output).resolve())
    pdf_paths = []
    skipped = []
    for p in all_pdfs:
        if str(Path(p).resolve()) == output_resolved:
            skipped.append(Path(p).name + " (output file)")
            continue
        if is_already_annotated(p):
            skipped.append(Path(p).name + " (already annotated)")
            continue
        pdf_paths.append(p)
    if skipped:
        print(f"Skipping {len(skipped)} file(s) that are already annotated or are the output: {skipped}")
    if not pdf_paths:
        sys.exit(f"ERROR: No unannotated source PDF files found in {spec_dir}")
    print(f"Found {len(pdf_paths)} source PDF file(s): {[Path(p).name for p in pdf_paths]}")

    # Parse manual map if provided
    manual_map = json.loads(args.manual_map) if args.manual_map else None

    # Read schedule
    print(f"\nReading schedule: {schedule_path.name}")
    schedule_entries = read_schedule(str(schedule_path), args.schedule_sheet)
    if not schedule_entries:
        sys.exit("ERROR: No valid entries found in schedule. Check CODE and MODEL columns.")
    print(f"  {len(schedule_entries)} products found in schedule.")

    # Map pages to codes
    print("\nMapping spec sheet pages to schedule codes…")
    page_code_map = build_page_code_map(pdf_paths, schedule_entries, manual_map)

    # Check for unmatched schedule entries
    matched_codes = set(page_code_map.values())
    unmatched = [(code, tokens[0]) for code, tokens in schedule_entries
                 if code not in matched_codes]
    if unmatched:
        print("\nWARNING: No spec sheet pages found for these schedule entries:")
        for code, model in unmatched:
            print(f"  {code} ({model}) — check that this product has a spec PDF in {spec_dir}")

    # Build annotated output
    print(f"\nAnnotating pages and building output PDF…")
    build_output_pdf(pdf_paths, page_code_map, schedule_entries, args.order, args.output)


if __name__ == "__main__":
    main()
